import threading
import time
import cv2
from ultralytics import YOLO
from datetime import datetime
from openpyxl import Workbook
import os

# 载入 YOLOv8 模型
model = YOLO('../best.pt')

camera_urls = [
    'rtsp://admin:admin@10.250.6.206/1',
    'rtsp://admin:admin@10.250.6.209/1',
    'rtsp://admin:admin@10.250.6.212/1',
    'rtsp://admin:admin@10.250.6.215/1',
    'rtsp://admin:admin@10.250.6.218/1',
    'rtsp://admin:admin@10.250.6.224/1',
    'rtsp://admin:admin@10.250.6.227/1',
    'rtsp://admin:admin@10.250.6.230/1',
    'rtsp://admin:admin@10.250.6.233/1',
    'rtsp://admin:admin@10.250.6.236/1',
    'rtsp://admin:admin@10.250.6.239/1',
    'rtsp://admin:admin@10.250.6.242/1',
    'rtsp://admin:admin@10.250.6.245/1',
    'rtsp://admin:admin@10.250.6.248/1',
    'rtsp://admin:admin@10.250.6.251/1',
    'rtsp://admin:admin@10.250.7.2/1',
    'rtsp://admin:admin@10.250.7.5/1',
    'rtsp://admin:admin@10.250.7.8/1',
    'rtsp://admin:admin@10.250.7.11/1',
    'rtsp://admin:admin@10.250.7.14/1',
    'rtsp://admin:admin@10.250.7.17/1',
    'rtsp://admin:admin@10.250.7.20/1',

    # 添加更多摄像头 URL ...
]

# 使用时间戳创建输出文件名
output_file = f"detection_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
temp_file = f"temp_{output_file}"  # 临时文件名

# 创建一个新的 Workbook
wb = Workbook()
ws = wb.active
ws.title = "Detection Results"

# 设置列名
columns = ["Camera URL", "Timestamp 1", "IsHavePeople 1", "PeopleCount 1"]
ws.append(columns)

# 保存初始文件
wb.save(temp_file)


# 定义写入 Excel 的函数
def write_to_excel(camera_url, current_time, IsHavePeople, people_count, round_num, max_retries=5, retry_delay=5):
    global wb, ws, output_file, temp_file
    retries = 0
    while retries < max_retries:
        try:
            # 找到相应的行
            existing_rows = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row + 1)]

            if camera_url not in existing_rows:
                # 如果相机 URL 不存在，则直接在新的一行添加数据
                new_row = [camera_url, current_time, IsHavePeople, people_count]
                ws.append(new_row)
            else:
                # 如果相机 URL 已存在，获取行索引
                camera_row_index = existing_rows.index(camera_url) + 2  # +2 to account for the header row

                # 在新列中添加数据
                ws.cell(row=camera_row_index, column=5 + (round_num - 1) * 3 + 1, value=current_time)  # Timestamp
                ws.cell(row=camera_row_index, column=6 + (round_num - 1) * 3 + 1, value=IsHavePeople)  # IsHavePeople
                ws.cell(row=camera_row_index, column=7 + (round_num - 1) * 3 + 1, value=people_count)  # PeopleCount

            # 保存文件到临时文件
            wb.save(temp_file)

            # 将临时文件重命名为目标文件
            if os.path.exists(output_file):
                os.remove(output_file)  # 删除旧的输出文件
            os.rename(temp_file, output_file)  # 重命名为最终文件
            break  # 如果保存成功，退出重试循环
        except PermissionError as e:
            retries += 1
            print(f"PermissionError: {e}. Retrying in {retry_delay} seconds...")
            time.sleep(retry_delay)
    else:
        print(f"Failed to save {output_file} after {max_retries} retries.")


# 定义检测函数
def detect_people(camera_url):
    round_num = 0
    cap = cv2.VideoCapture(camera_url)
    if not cap.isOpened():
        print(f"Error: Could not open video stream from {camera_url}")
        return

    while True:
        ret, frame = cap.read()
        if not ret:
            print(f"Error: Could not read frame from {camera_url}")
            break

        # 检测物体
        results = model(frame)

        if results and hasattr(results[0], 'boxes'):
            boxes = results[0].boxes
            people_count = len(boxes)
            IsHavePeople = "是" if people_count > 0 else "否"
        else:
            people_count = 0
            IsHavePeople = "否"

        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # 写入检测结果到 Excel
        print(f"Writing data to Excel for {camera_url}: {current_time}, {IsHavePeople}, {people_count}")
        write_to_excel(camera_url, current_time, IsHavePeople, people_count, round_num)

        # 每隔10秒检测一次
        time.sleep(30)

        # 在每次检测后更新轮次
        round_num += 1

    cap.release()


# 创建并启动线程
threads = []
for url in camera_urls:
    t = threading.Thread(target=detect_people, args=(url,))
    t.start()
    threads.append(t)

# 等待所有线程完成
for t in threads:
    t.join()
